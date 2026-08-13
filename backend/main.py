import os
import io
import json
import asyncio
import pandas as pd
import requests
from fastapi import FastAPI, File, UploadFile, Form
from fastapi.responses import JSONResponse
from fastapi.staticfiles import StaticFiles
from playwright.async_api import async_playwright
import uuid
import datetime
import openpyxl
from openpyxl.styles import PatternFill

import sys
import logging

# Configure production logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger("subsidiaries-ai")

# Setup directories
DOWNLOADS_DIR = os.path.join(os.path.dirname(__file__), "downloads")
os.makedirs(DOWNLOADS_DIR, exist_ok=True)
HISTORY_FILE = os.path.join(os.path.dirname(__file__), "audit_history.json")

import sqlite3

def load_history():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, "r") as f:
            return json.load(f)
    return []

def save_history(history):
    with open(HISTORY_FILE, "w") as f:
        json.dump(history, f, indent=4)

# SQLite Caching Setup
CACHE_DB = os.path.join(os.path.dirname(__file__), "cache.db")
def init_db():
    conn = sqlite3.connect(CACHE_DB)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS scan_cache (
            parent TEXT,
            entity TEXT,
            result_json TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (parent, entity)
        )
    ''')
    conn.commit()
    conn.close()

init_db()

def get_cached_result(parent, entity):
    conn = sqlite3.connect(CACHE_DB)
    cursor = conn.cursor()
    cursor.execute('SELECT result_json FROM scan_cache WHERE parent=? AND entity=?', (parent, entity))
    row = cursor.fetchone()
    conn.close()
    if row:
        return json.loads(row[0])
    return None

def set_cached_result(parent, entity, result_dict):
    conn = sqlite3.connect(CACHE_DB)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT OR REPLACE INTO scan_cache (parent, entity, result_json, timestamp)
        VALUES (?, ?, ?, CURRENT_TIMESTAMP)
    ''', (parent, entity, json.dumps(result_dict)))
    conn.commit()
    conn.close()


# Fix for Windows console unicode printing errors
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from fastapi import FastAPI, WebSocket, WebSocketDisconnect

app = FastAPI()

class ConnectionManager:
    def __init__(self):
        self.active_connections: dict[str, WebSocket] = {}

    async def connect(self, websocket: WebSocket, client_id: str):
        await websocket.accept()
        self.active_connections[client_id] = websocket

    def disconnect(self, client_id: str):
        if client_id in self.active_connections:
            del self.active_connections[client_id]

    async def send_personal_message(self, message: dict, client_id: str):
        if client_id and client_id in self.active_connections:
            try:
                await self.active_connections[client_id].send_json(message)
            except Exception as e:
                pass

manager = ConnectionManager()

@app.websocket("/ws/progress/{client_id}")
async def websocket_endpoint(websocket: WebSocket, client_id: str):
    await manager.connect(websocket, client_id)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(client_id)


from bs4 import BeautifulSoup
from playwright.async_api import async_playwright

playwright_manager = None
browser_instance = None
browser_context = None
browser_sem = asyncio.Semaphore(3)  # Maximum 3 concurrent tabs

@app.on_event("startup")
async def startup_event():
    global playwright_manager, browser_instance, browser_context
    try:
        playwright_manager = await async_playwright().start()
        browser_instance = await playwright_manager.chromium.launch(headless=True)
        browser_context = await browser_instance.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
        )
        logger.info("Playwright browser started.")
    except Exception as e:
        logger.error(f"Failed to start Playwright: {e}")

@app.on_event("shutdown")
async def shutdown_event():
    global playwright_manager, browser_instance, browser_context
    if browser_context:
        await browser_context.close()
    if browser_instance:
        await browser_instance.close()
    if playwright_manager:
        await playwright_manager.stop()
    logger.info("Playwright browser stopped.")

from fastapi.middleware.cors import CORSMiddleware

# Enable CORS for GitHub Pages
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://tarun1790.github.io", "http://localhost:8000", "http://127.0.0.1:8000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

import os
# Ensure frontend directory exists for static files to mount successfully
frontend_dir = os.path.join(os.path.dirname(__file__), "../docs")

from fastapi.responses import FileResponse
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()



GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
SERPER_API_KEY = os.getenv("SERPER_API_KEY", "")

import asyncio

def _fetch_serper(query: str) -> list:
    """Synchronous function to hit Serper.dev and extract snippets and links."""
    if not SERPER_API_KEY:
        return []
        
    url = "https://google.serper.dev/search"
    payload = json.dumps({"q": query, "num": 2})
    headers = {
        'X-API-KEY': SERPER_API_KEY,
        'Content-Type': 'application/json'
    }
    
    import time
    
    try:
        # Retry up to 3 times for 429 Rate Limits
        for attempt in range(4):
            response = requests.post(url, headers=headers, data=payload, timeout=15)
            if response.status_code == 429:
                if attempt < 3:
                    time.sleep(1 + attempt)  # Backoff
                    continue
                else:
                    return []
                    
            if response.status_code != 200:
                return []
                
            data = response.json()
            results = []
            
            # Extract Google Answer Box (AI Overview / Direct Answer)
            if "answerBox" in data:
                ans = data["answerBox"]
                if "snippet" in ans:
                    results.append({"snippet": f"Google Answer Box: {ans['snippet']}", "link": ans.get("link", "")})
                elif "answer" in ans:
                    results.append({"snippet": f"Google Answer Box: {ans['answer']}", "link": ans.get("link", "")})
            
            # Extract Google Knowledge Graph
            if "knowledgeGraph" in data:
                kg = data["knowledgeGraph"]
                kg_text = f"Google Knowledge Graph: Title: {kg.get('title', '')} | Type: {kg.get('type', '')} | Desc: {kg.get('description', '')}"
                attrs = kg.get("attributes", {})
                if isinstance(attrs, dict):
                    for k, v in attrs.items():
                        kg_text += f" | {k}: {v}"
                results.append({"snippet": kg_text, "link": kg.get("descriptionLink", "")})

            for result in data.get("organic", [])[:2]:
                if "link" in result:
                    results.append({
                        "snippet": result.get("snippet", ""),
                        "link": result.get("link", "")
                    })
            return results
    except Exception as e:
        logger.error(f"Serper error: {e}")
        return []

async def scrape_url(url: str) -> str:
    global browser_context
    if not browser_context:
        return ""
        
    async with browser_sem:
        page = None
        try:
            page = await browser_context.new_page()
            # Abort media requests to speed up load and save memory
            await page.route("**/*", lambda route: route.abort() if route.request.resource_type in ["image", "media", "font", "stylesheet"] else route.continue_())
            await page.goto(url, timeout=5000, wait_until="domcontentloaded")
            html = await page.content()
            soup = BeautifulSoup(html, 'lxml')
            
            # Remove scripts, styles, navs
            for script in soup(["script", "style", "nav", "footer", "header", "aside"]):
                script.decompose()
            
            text = soup.get_text(separator=' ', strip=True)
            # Limit text size to ~5000 characters
            return text[:5000]
        except Exception as e:
            logger.warning(f"Scrape failed for {url}: {e}")
            return ""
        finally:
            if page:
                await page.close()

async def fetch_search_data(entity: str, parent: str, sem) -> tuple[str, str]:
    """Uses Serper.dev and Playwright to fetch full web evidence."""
    async with sem:
        query = f"{entity} {parent}"
        results = await asyncio.to_thread(_fetch_serper, query)
        
        evidence_parts = []
        for res in results:
            evidence = f"Snippet: {res['snippet']}"
            if res['link']:
                scraped_text = await scrape_url(res['link'])
                if scraped_text:
                    evidence += f"\nFull Page Text ({res['link']}): {scraped_text}"
            evidence_parts.append(evidence)
            
        final_evidence = "\n\n".join(evidence_parts) if evidence_parts else "No snippets found."
        return entity, final_evidence

def analyze_relationships_batch(entities: list, parent: str, search_results: dict) -> tuple[list, int]:
    """Uses Gemini API to evaluate a batch of entities based on internal AI knowledge."""
    
    # We construct a numbered list of entities with their corresponding web search evidence
    entities_list_str = ""
    for i, ent in enumerate(entities):
        entities_list_str += f"{i+1}. {ent}\n"
        entities_list_str += f"   Web Evidence: {search_results.get(ent, 'No data')}\n\n"
    
    prompt = f"""
    You are a corporate relationship analyzer.
    I will give you a parent company and a list of entities along with recent web search evidence for each.
    
    Parent Company: '{parent}'
    
    Entities to check:
    {entities_list_str}
    
    "Is Entity X actually related to Parent Company Y?"

    Do NOT answer this based on the entity name.
    Do NOT assume ownership because:
    - The companies operate in the same industry.
    - They operate in the same country.
    - Their names are similar.
    - One company mentions the other.
    - They have a commercial partnership.
    - They distribute each other's products.
    - They have customers or suppliers in common.
    - They appear together in search results.

    A corporate relationship must be established by evidence.

    RELATIONSHIP TYPES
    Determine the actual relationship from the following categories:

    1. DIRECT_SUBSIDIARY: The specified parent directly owns or controls the entity.
    2. INDIRECT_SUBSIDIARY: The specified parent ultimately controls the entity through one or more intermediate subsidiaries.
    3. SISTER_COMPANY: Both are controlled by the same ultimate parent, but the specified parent does not control the entity. (NOT a subsidiary).
    4. PARENT_COMPANY: The entity is actually the parent of the specified company. (NOT a subsidiary).
    5. BRAND: The entity is a brand or trademark rather than a separate legal subsidiary. (NOT a subsidiary automatically).
    6. BUSINESS_UNIT: An internal division rather than a separate legal entity.
    7. JOINT_VENTURE: Jointly controlled or owned by multiple parties.
    8. ASSOCIATED_COMPANY: The parent has an investment, significant influence, or strategic relationship.
    9. FORMER_SUBSIDIARY: Previously owned, but no longer controlled. (NOT a current subsidiary).
    10. ACQUIRED_COMPANY: Acquired and currently controlled by the parent. (Do not treat announced/proposed as completed).
    11. DIVESTED_COMPANY: Parent previously owned but sold it. (NOT a subsidiary).
    12. UNRELATED: No established corporate ownership or control relationship.
    13. UNKNOWN: The available evidence is insufficient to determine the relationship. Do NOT guess.

    VERIFICATION RULES
    RULE 1 — EXTREME BENEFIT OF THE DOUBT: You MUST assume the provided spreadsheet is generally correct. If there is ANY conceivable link, you MUST classify it as a MATCH. Only mark as NO MATCH if you can definitively prove the entity belongs entirely to a known competitor with zero ties. When in doubt, default to MATCH.
    RULE 2 — OWNERSHIP IS DIFFERENT FROM EXISTENCE: Existence-only evidence is insufficient to prove ownership, but refer to Rule 1.
    RULE 3 — ESTABLISH RELATIONSHIP: You do NOT need official documents (like SEC filings). Use your understanding of the web search snippets.
    RULE 4 — CURRENT RELATIONSHIP: Determine whether the relationship exists CURRENTLY.
    RULE 5 — INDIRECT OWNERSHIP: Follow the ownership chain when evidence supports it.
    RULE 6 — SISTER COMPANIES: Do not confuse common ownership with parent-child ownership.
    RULE 7 — BRAND ≠ LEGAL ENTITY: A brand name does not automatically mean it is a legal subsidiary, but it STILL counts as a MATCH.
    RULE 8 — ACQUISITION STATUS: Only completed control should establish current ownership.
    RULE 9 — LOGICAL INFERENCE: If the web evidence is sparse, you MUST logically deduce if the entity is related to the parent based on context.
    RULE 10 — LOCATION AND INDUSTRY: Same country/industry alone does not establish ownership, but refer to Rule 1.
    RULE 11 — INTERNAL CORPORATE KNOWLEDGE: You MUST use your own internal AI knowledge of corporate history, acquisitions, and groups. Do NOT reject an entity just because the provided web snippets are sparse if your internal knowledge confirms the relationship.

    ECOSYSTEM INCLUSION RULES (RELAXED VERIFICATION)
    - If the entity is a generic topic, subject, or service that the parent company teaches or provides (e.g., 'Minority languages' for an education company), it MUST BE CLASSIFIED AS A MATCH (CORRECT).
    - If the entity is an internal business division, department, or operational center, it MUST BE CLASSIFIED AS A MATCH (CORRECT).
    - If the entity is a product, app, digital wallet, portal, service, or brand that is owned, developed, or operated by the parent company OR by ANY of its subsidiaries, it MUST BE CLASSIFIED AS A MATCH (CORRECT).
    - If the entity is a mutual fund, REIT, private equity fund, or any other financial fund managed or launched by the parent, its subsidiaries, or its founders, it MUST BE CLASSIFIED AS A MATCH (CORRECT).
    - If the parent company or a child of the parent company has invested funds into the entity, it MUST BE CLASSIFIED AS A MATCH (CORRECT).
    - In short: if an entity is part of the parent's ecosystem directly or indirectly, or if there is investment by the parent/subsidiary, it is a MATCH. 
    - Only mark as NO MATCH (INCORRECT) if the entity is completely independently owned by a totally unrelated party or competitor.

    EVIDENCE SUFFICIENCY
    Before declaring a relationship, ask:
    - Does the evidence identify the entity and parent?
    - Does the evidence establish ecosystem inclusion, ownership, investment, or control?
    - Can a relationship to the parent or a superparent be logically deduced from the entity name and context?
    If there is absolutely zero evidence and no logical inference can be made, the relationship is UNKNOWN. However, if a relationship can be reasonably inferred, classify it as a MATCH.

    DECISION LOGIC
    Determine:
    A. VERIFIED RELATIONSHIP: What is the entity's actual relationship?
    B. VERIFIED PARENT: Who actually owns, operates, or invested in the entity?
    C. CLAIMED PARENT: What parent company does the spreadsheet claim?
    D. MATCH: Does the verified relationship support the claimed parent relationship? (Direct, Indirect, Joint Venture, Associated, Acquired, Ecosystem Product/Fund/Investment, Superparent/Holding Relationship, or Logically Inferred = MATCH. Unrelated, Sister, Former, Unknown = NO MATCH).

    Respond strictly in JSON format. The JSON should be an array of objects for EVERY entity provided in the chunk (both matches and non-matches).
    
    Format:
    [
      {{
        "name": "Exact Entity Name from the list",
        "is_match": true or false,
        "confidence": "high, medium, low, or none",
        "evidence": "Briefly state the specific evidence or internal knowledge cited (e.g. 'Google Knowledge Graph states 100% owned', or 'Internal knowledge confirms it is a division of XYZ').",
        "reason": "Detailed explanation of why it is or isn't a match. Do not use boilerplate text. If not a match, name the ACTUAL parent or owner."
      }}
    ]
    
    Do not skip any entities. Every entity in the chunk must have a corresponding object in the JSON array.
    Do not output any markdown formatting, only pure JSON.
    """
    
    models_to_try = [
        "gemini-3.6-flash", 
        "gemini-3.5-flash-lite", 
        "gemini-3.0-flash", 
        "gemini-3.1-flash-lite"
    ]
    
    for model_name in models_to_try:
        try:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent?key={GEMINI_API_KEY}"
            response = requests.post(
                url, 
                headers={
                    "Content-Type": "application/json"
                },
                data=json.dumps({
                    "contents": [{
                        "parts": [{"text": prompt}]
                    }],
                    "generationConfig": {
                        "temperature": 0.2,
                        "responseMimeType": "application/json"
                    }
                }),
                timeout=120
            )
            
            if response.status_code == 429:
                logger.info(f"Rate limited on {model_name}, falling back to next model...")
                continue
                
            if response.status_code != 200:
                logger.error(f"API Error {response.status_code} on {model_name}, falling back...")
                continue
            
            data = response.json()
            tokens_used = 0
            if 'usageMetadata' in data:
                tokens_used = data['usageMetadata'].get('totalTokenCount', 0)
                
            if 'candidates' in data and len(data['candidates']) > 0:
                content = data['candidates'][0]['content']['parts'][0].get('text', '').strip()
                # Clean up markdown if any
                if content.startswith('```json'):
                    content = content[7:]
                if content.startswith('```'):
                    content = content[3:]
                if content.endswith('```'):
                    content = content[:-3]
                    
                return json.loads(content), tokens_used
        except Exception as e:
            logger.error(f"Batch API exception on {model_name}: {e}")
            continue
            
    return [{"name": "Error", "reason": "Failed to analyze batch due to API error across all models."}], 0

@app.post("/api/verify")
async def verify_subsidiaries(parent_name: str = Form(...), file: UploadFile = File(...), client_id: str = Form(None)):
    try:
        contents = await file.read()
        
        # Load the workbook with openpyxl to preserve formatting
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        
        # Look for a specific sheet containing "legalentitymapping" or "lem"
        target_sheet_name = None
        for sheet_name in wb.sheetnames:
            clean_name = sheet_name.lower().replace(' ', '')
            if 'legalentitymapping' in clean_name or 'legalenititymapping' in clean_name or 'lem' in clean_name:
                target_sheet_name = sheet_name
                break
                
        if target_sheet_name:
            ws = wb[target_sheet_name]
        else:
            ws = wb.active
        
        # Dynamically detect which column contains the entity name
        entity_col_idx = 1
        header_keywords = ['entity', 'lem', 'company', 'legalenititymapping', 'legalentitymapping', 'account', 'accountname', 'subsidiariesinqa']
        
        for col in range(1, ws.max_column + 1):
            header_val = ws.cell(row=1, column=col).value
            if header_val:
                header_str = str(header_val).lower().replace(' ', '')
                # Skip any column that is an ID column (e.g. "Account ID", "Entity ID")
                if 'id' in header_str:
                    continue
                if any(kw in header_str for kw in header_keywords):
                    entity_col_idx = col
                    break
        
        # Extract entities from the detected column, skipping the header (row 1)
        raw_entities = []
        for row in ws.iter_rows(min_row=2, min_col=entity_col_idx, max_col=entity_col_idx, values_only=True):
            if row[0]:
                raw_entities.append(str(row[0]))
                
        # Additional filter just in case
        entities = [e for e in raw_entities if e.lower().strip() not in ['entity name', 'lem', 'company name', 'entity']]
        
        if not entities:
            return JSONResponse(content={"incorrect_lems": [], "total_checked": 0})
            
        # 1. Check Cache First
        if client_id: await manager.send_personal_message({"step": "Checking Cache...", "progress": 10}, client_id)
        all_lems = []
        uncached_entities = []
        for entity in entities:
            cached = get_cached_result(parent_name, entity)
            if cached:
                all_lems.append(cached)
            else:
                uncached_entities.append(entity)
                
        logger.info(f"{len(entities) - len(uncached_entities)} entities loaded from cache. {len(uncached_entities)} require web search.")
        
        total_tokens_used = 0
        
        if uncached_entities:
            if client_id: await manager.send_personal_message({"step": f"Searching Web for {len(uncached_entities)} entities...", "progress": 30}, client_id)
            logger.info(f"Searching web concurrently via Serper for {len(uncached_entities)} entities...")
            search_results = {}
            sem = asyncio.Semaphore(4) # Serper Free tier allows lower QPS, reduced to 4
            tasks = [fetch_search_data(entity, parent_name, sem) for entity in uncached_entities]
            results = await asyncio.gather(*tasks)
            
            for ent, data in results:
                search_results[ent] = data
                
            logger.info("Scraping completed. Analyzing with Gemini in chunks of 50...")
            if client_id: await manager.send_personal_message({"step": "Analyzing web data with Gemini...", "progress": 50}, client_id)
            
            chunk_size = 50
            for i in range(0, len(uncached_entities), chunk_size):
                chunk = uncached_entities[i:i + chunk_size]
                logger.info(f"Sending chunk {i // chunk_size + 1} to Gemini...")
                chunk_results, tokens = analyze_relationships_batch(chunk, parent_name, search_results)
                total_tokens_used += tokens
                
                # First pass
                low_confidence_entities = []
                for item in chunk_results:
                    if item.get("name") != "Error":
                        confidence = item.get("confidence", "").lower()
                        is_match = item.get("is_match")
                        ent_name = item.get('name', item.get('entity', item.get('company', '')))
                        
                        if confidence in ["low", "none"] and not is_match and ent_name:
                            low_confidence_entities.append(ent_name)
                        else:
                            all_lems.append(item)
                            if ent_name:
                                set_cached_result(parent_name, ent_name, item)
                
                # Second pass (Deep Research Fallback)
                if low_confidence_entities:
                    if client_id: await manager.send_personal_message({"step": f"Deep Research Fallback for {len(low_confidence_entities)} entities...", "progress": 75}, client_id)
                    logger.info(f"Triggering Deep Research Fallback for {len(low_confidence_entities)} low-confidence entities...")
                    deep_search_results = {}
                    sem2 = asyncio.Semaphore(4)
                    
                    async def fetch_deep(ent, parent, semaphore):
                        async with semaphore:
                            deep_query = f'"{ent}" AND ("subsidiary" OR "acquired" OR "owned by") "{parent}"'
                            res = await asyncio.to_thread(_fetch_serper, deep_query)
                            if not res:
                                deep_query_2 = f'"{ent}" parent company owner'
                                res = await asyncio.to_thread(_fetch_serper, deep_query_2)
                            return (ent, res)
                            
                    deep_tasks = [fetch_deep(e, parent_name, sem2) for e in low_confidence_entities]
                    deep_res = await asyncio.gather(*deep_tasks)
                    
                    for ent, data in deep_res:
                        deep_search_results[ent] = data
                        
                    logger.info(f"Deep Search completed. Re-analyzing {len(low_confidence_entities)} entities...")
                    deep_chunk_results, deep_tokens = analyze_relationships_batch(low_confidence_entities, parent_name, deep_search_results)
                    total_tokens_used += deep_tokens
                    
                    for item in deep_chunk_results:
                        if item.get("name") != "Error":
                            all_lems.append(item)
                            ent_name = item.get('name', item.get('entity', item.get('company', '')))
                            if ent_name:
                                set_cached_result(parent_name, ent_name, item)
                    
        if client_id: await manager.send_personal_message({"step": "Generating Excel report...", "progress": 90}, client_id)
        # Highlight entities in the Excel file and add new columns
        red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        green_fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")
        
        # Build a map from lowercase name to the result object
        result_map = {}
        for item in all_lems:
            ent_name = item.get('name', item.get('entity', item.get('company', '')))
            if ent_name:
                result_map[ent_name.lower()] = item
        
        # Find the next empty column to append details without overwriting existing data
        start_col_idx = ws.max_column + 1
        
        # Add headers for the new columns
        ws.cell(row=1, column=start_col_idx, value="Match Status")
        ws.cell(row=1, column=start_col_idx + 1, value="Confidence")
        ws.cell(row=1, column=start_col_idx + 2, value="Evidence")
        ws.cell(row=1, column=start_col_idx + 3, value="Reason")
        
        for row in ws.iter_rows(min_row=2, min_col=entity_col_idx, max_col=entity_col_idx):
            cell = row[0]
            if cell.value:
                val_lower = str(cell.value).lower()
                if val_lower in result_map:
                    res = result_map[val_lower]
                    is_match = res.get('is_match', False)
                    cell.fill = green_fill if is_match else red_fill
                    
                    # Write details in the new columns
                    ws.cell(row=cell.row, column=start_col_idx).value = "MATCH" if is_match else "NO MATCH"
                    ws.cell(row=cell.row, column=start_col_idx + 1).value = str(res.get('confidence', 'none')).upper()
                    ws.cell(row=cell.row, column=start_col_idx + 2).value = res.get('evidence', '')
                    ws.cell(row=cell.row, column=start_col_idx + 3).value = res.get('reason', '')
                
        run_id = str(uuid.uuid4())
        output_filename = f"marked_{run_id}.xlsx"
        output_path = os.path.join(DOWNLOADS_DIR, output_filename)
        wb.save(output_path)
        
        incorrect_count = sum(1 for x in all_lems if not x.get('is_match', False))
        incorrect_lems = [x for x in all_lems if not x.get('is_match', False)]

        # Log to history
        history = load_history()
        history.insert(0, {
            "id": run_id,
            "timestamp": datetime.datetime.now().isoformat(),
            "parent_name": parent_name,
            "filename": file.filename,
            "total_checked": len(entities),
            "incorrect_count": incorrect_count,
            "tokens_used": total_tokens_used,
            "incorrect_lems": incorrect_lems, # Keep for backwards compatibility
            "all_lems": all_lems  # Save the full detailed scan data
        })
        save_history(history)
                
        if client_id: await manager.send_personal_message({"step": "Complete", "progress": 100}, client_id)
        
        return JSONResponse(content={
            "all_lems": all_lems,
            "incorrect_lems": incorrect_lems, 
            "total_checked": len(entities),
            "download_url": f"/api/download/{run_id}",
            "run_id": run_id,
            "tokens_used": total_tokens_used
        })
        
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)

@app.get("/api/history")
async def get_history():
    return JSONResponse(content=load_history())

@app.get("/api/download/{run_id}")
async def download_file(run_id: str):
    file_path = os.path.join(DOWNLOADS_DIR, f"marked_{run_id}.xlsx")
    if os.path.exists(file_path):
        history = load_history()
        run_data = next((item for item in history if item.get("id") == run_id), None)
        
        # Use the parent company name as the filename
        if run_data and run_data.get("parent_name"):
            # Sanitize the parent name to be a valid filename (replace spaces and special chars)
            clean_parent = "".join(c if c.isalnum() else "_" for c in run_data["parent_name"])
            download_name = f"{clean_parent}.xlsx"
        else:
            download_name = "Verified_Results.xlsx"
            
        return FileResponse(
            file_path, 
            filename=download_name, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
                "Pragma": "no-cache",
                "Expires": "0"
            }
        )
    return JSONResponse(content={"error": "File not found"}, status_code=404)

if os.path.exists(frontend_dir):
    app.mount("/", StaticFiles(directory=frontend_dir, html=True), name="docs")

if __name__ == "__main__":
    import uvicorn
    import asyncio
    if sys.platform == 'win32':
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)
